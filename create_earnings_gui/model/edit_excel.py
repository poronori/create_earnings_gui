import openpyxl
import configparser
import datetime
from copy import copy
from .scraping_data import ScrapingData

def edit(dataList) : 
    print('===========エクセルに記入 開始===========')
    
    #設定ファイルからエクセルのパスを取得
    inifile = configparser.SafeConfigParser()
    inifile.read('config/config.ini', encoding='utf-8')
    excel_path = inifile.get('DEFAULT', 'ExcelPath')
    print(excel_path)
    wb = openpyxl.load_workbook(excel_path)
    
    #販売リスト
    for data in dataList:
        add_create_list(wb, data)
    #宛名
    ws = wb['宛名']
    edit_addressee_atena(ws, dataList)
    ws = wb['宛名 圧迫厳禁']
    edit_addressee_appaku(ws, dataList)

    wb.save(excel_path)
    
    print('===========エクセルに記入 終了===========')

# 販売リストに追記
def add_create_list(wb, data:ScrapingData):
    
    date = data.get_date()
    name = data.get_name()
    price = data.get_price()
    commission = data.get_commission()
    customer = data.get_customer()
    address = data.get_address()
    code = data.get_code()
    current_year = data.date.today().year
    
    print('購入日：   ' + date)
    print('品名：     ' + name)
    print('商品代金：  ' + str(price))
    print('販売手数料：' + str(commission))
    print('購入者：    ' + customer)
    print('配送先；    ' + address)
    print('コード：    ' + code)
    
    ws = wb['販売リスト %s' %current_year]
    
    #一番下の行を取得
    maxRow = ws.max_row
    #max_rowを使うと削除していた行も取得してしまうため、Noneで判定する
    for row in ws.iter_rows():
        if not row[0] or row[0].value is None :
            maxRow = row[0].row - 1
            break
    print('追加行:' + str(maxRow))
    nextRow = maxRow + 1
    
    #行を挿入
    ws.insert_rows(nextRow)
    
    #書式をコピー
    i = 1
    profit = ''
    for row in ws.iter_rows():
        for cell in row:
            if cell.row == maxRow:
                ws.cell(row = nextRow, column = i).border = copy(cell.border)
                ws.cell(row = nextRow, column = i)._style = copy(cell._style)
                value = cell.value
                #計算式をコピー
                if str(value)[0] == "=":
                    profit = value.replace(str(maxRow), str(nextRow))
                i = i + 1

    #値をセット
    no = ws.cell(row = maxRow, column = 1).value
    ws.cell(row = nextRow, column = 1).value = int(no) + 1   #No
    ws.cell(row = nextRow, column = 2).value = date          #購入日
    ws.cell(row = nextRow, column = 3).value = name          #品名
    ws.cell(row = nextRow, column = 4).value = price         #商品代金
    ws.cell(row = nextRow, column = 5).value = commission    #販売手数料
    ws.cell(row = nextRow, column = 6).value = ''            #梱包資材１   
    ws.cell(row = nextRow, column = 7).value = '封筒'        #梱包資材２
    ws.cell(row = nextRow, column = 8).value = ''            #送料
    ws.cell(row = nextRow, column = 9).value = profit        #販売利益
    ws.cell(row = nextRow, column = 10).value = customer     #購入者
    ws.cell(row = nextRow, column = 11).value = address      #住所
    ws.cell(row = nextRow, column = 12).value = code         #コード

# 宛名シート
def edit_addressee_atena(ws, dataList):
    
    # 1件目のみ挿入
    data = dataList[0]
    postcode = data.postcode
    address1 = data.get_address1()
    address2 = data.get_address2()
    customer = data.get_customer_full()
    
    ws.cell(row = 2, column = 2).value = postcode
    ws.cell(row = 3, column = 2).value = address1
    ws.cell(row = 4, column = 2).value = address2
    ws.cell(row = 6, column = 2).value = customer

# 宛名 圧迫厳禁シート
def edit_addressee_appaku(ws, dataList):
    
    i = 1
    column = 2
    row = 2
    
    for data in dataList:
        
        # 4件目まで挿入
        while i <= 4:
            postcode = data.postcode
            address1 = data.get_address1()
            address2 = data.get_address2()
            customer = data.get_customer_full()

            ws.cell(row = row, column = column).value = postcode
            ws.cell(row = row + 1, column = column).value = address1
            ws.cell(row = row + 2, column = column).value = address2
            ws.cell(row = row + 4, column = column).value = customer
            
            #2件目と4件目は下に、3件目は右上に移動
            if i % 2 == 0:
                row -= 12
                column += 4
            else:
                row += 12
            i += 1

def edit_addressee(ws, dataList):
    
    i = 1
    column = 2
    row = 2
    
    for data in dataList:
        
        #3件目以降は書式をコピーする
        if i > 1:
            #基準は左上のセル
            for base_row in range(2, 12):
                for base_column in range(2, 4):
                    #偶数件目なら右側に追加
                    if i % 2 == 0:
                        ws.cell(row = base_row + row - 2, column = base_column + 4).border = copy(ws.cell(row = base_row, column = base_column).border)
                        ws.cell(row = base_row + row - 2, column = base_column + 4)._style = copy(ws.cell(row = base_row, column = base_column)._style)
                    #奇数件目なら左下に追加
                    else:
                        ws.cell(row = base_row + row - 2, column = base_column).border = copy(ws.cell(row = base_row, column = base_column).border)
                        ws.cell(row = base_row + row - 2, column = base_column)._style = copy(ws.cell(row = base_row, column = base_column)._style)
            
            #自宅アドレスのセルは値もコピーする
            for base_row in range(8, 12):
                if i % 2 == 0:
                    ws.cell(row = base_row, column = base_column + 4).value = copy(ws.cell(row = base_row, column = base_column).border)
                else:
                    ws.cell(row = base_row + 12, column = base_column).border = copy(ws.cell(row = base_row, column = base_column).border)
            
        
        postcode = data.postcode
        address1 = data.get_address1()
        address2 = data.get_address2()
        customer = data.get_customer_full()
        
        ws.cell(row = row, column = column).value = postcode
        ws.cell(row = row + 1, column = column).value = address1
        ws.cell(row = row + 2, column = column).value = address2
        ws.cell(row = row + 4, column = column).value = customer
        
        #奇数件数の場合は横にずれる
        if i % 2 == 0:
            column += 4
        #偶数件数の場合は左下にずれる
        else:
            row += 12
            column -= 4
        i += 1

#日時を〇月×日にフォーマット
def string_to_datetime(date_string) :
    datetime_array = date_string.split('/')
    year = int(datetime_array[0])
    month = int(datetime_array[1])
    day = int(datetime_array[2])
    dt = datetime.date(year, month, day)
    
    return dt.strftime('%m月%d日')

#住所から県名を抽出
def prefecture_from_address(address) :
    index = address.find('県', 0, 4)
    if index != -1 :
        return address[0:index + 1]
    
    index = address.find('府', 0, 3)
    if index != -1 :
        return address[0:index + 1]
    
    if '北海道' in address :
        return address[0:3]
    
    if '東京都' in address :
        return address[0:3]